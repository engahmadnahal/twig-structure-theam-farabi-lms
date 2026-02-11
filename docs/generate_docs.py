#!/usr/bin/env python3
"""Generate Excel documentation for Farabi First Theme"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def style_header(ws, row, cols, fill_color="1e3a5f"):
    """Apply header styling"""
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_data(ws, start_row, end_row, cols, alternate=True):
    """Apply data row styling"""
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    light_fill = PatternFill(start_color='f0f7ec', end_color='f0f7ec', fill_type='solid')
    
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if alternate and (row - start_row) % 2 == 1:
                cell.fill = light_fill

def auto_width(ws, cols, min_width=12, max_width=50):
    """Auto-adjust column widths"""
    for col in range(1, cols + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3

def add_title(ws, title, row=1, cols=4):
    """Add a title row"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(name='Arial', bold=True, size=14, color='1e3a5f')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 35


# ============================================================
# FILE 1: Variables & Config
# ============================================================
wb1 = openpyxl.Workbook()

# --- Sheet 1: Twig Variables ---
ws1 = wb1.active
ws1.title = "Twig Variables"
ws1.sheet_properties.tabColor = "93c572"

add_title(ws1, "Twig Template Variables", 1, 4)

headers = ["Variable", "Type", "Description", "Default Value"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 4)

variables = [
    ["current_locale", "string", "اللغة الحالية - تتحكم بـ lang و dir", "'en'"],
    ["asset_path", "string", "مسار مجلد الأصول (CSS/JS)", "'' (فارغ)"],
    ["page_path", "string", "مسار الصفحات (للروابط بين الصفحات)", "'' (فارغ)"],
    ["active_page", "string", "اسم الصفحة النشطة لتمييزها في القائمة", "'' (فارغ)"],
    ["config", "object", "كائن إعدادات الثيم من config.json", "--"],
]
for i, row in enumerate(variables, 4):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)
style_data(ws1, 4, 4 + len(variables) - 1, 4)
auto_width(ws1, 4)

# --- Sheet 2: Config Keys ---
ws2 = wb1.create_sheet("Config Keys")
ws2.sheet_properties.tabColor = "1e3a5f"

add_title(ws2, "Config Object Keys (config.json)", 1, 4)

headers = ["Key", "Type", "Description (الوصف)", "Example (مثال)"]
for i, h in enumerate(headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 4)

config_keys = [
    ["config.site.name", "string", "اسم الموقع", "Farabi Academy"],
    ["config.site.logo_text", "string", "النص داخل اللوغو الدائري", "FA"],
    ["config.site.email", "string", "بريد التواصل", "info@farabiacademy.com"],
    ["config.site.phone", "string", "رقم الهاتف", "+1 (555) 123-4567"],
    ["config.site.address", "string", "العنوان", "123 Education Street, Learning City"],
    ["config.colors.pistachio", "string", "اللون الأساسي (أخضر فستقي)", "#93c572"],
    ["config.colors.pistachio_dark", "string", "اللون الأساسي الداكن", "#7db157"],
    ["config.colors.navy", "string", "اللون الثانوي (كحلي)", "#1e3a5f"],
    ["config.colors.navy_dark", "string", "اللون الثانوي الداكن", "#152a45"],
    ["config.social.facebook", "string", "رابط فيسبوك", "#"],
    ["config.social.instagram", "string", "رابط انستغرام", "#"],
    ["config.social.twitter", "string", "رابط تويتر", "#"],
    ["config.social.tiktok", "string", "رابط تيك توك", "#"],
]
for i, row in enumerate(config_keys, 4):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)
style_data(ws2, 4, 4 + len(config_keys) - 1, 4)
auto_width(ws2, 4)

output_path = os.path.join(os.path.dirname(__file__), "01-variables-config.xlsx")
wb1.save(output_path)
print(f"Created: {output_path}")


# ============================================================
# FILE 2: Twig Blocks
# ============================================================
wb2 = openpyxl.Workbook()
ws = wb2.active
ws.title = "Twig Blocks"
ws.sheet_properties.tabColor = "93c572"

add_title(ws, "Twig Blocks (بلوكات قابلة للتعديل)", 1, 4)

headers = ["Block Name", "Location (الموقع)", "Description (الوصف)", "Used In (مُستخدم في)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 4)

blocks = [
    ["{% block title %}", "<title>", "عنوان الصفحة في التبويب", "كل الصفحات"],
    ["{% block meta_description %}", "<meta>", "وصف الصفحة لـ SEO", "كل الصفحات"],
    ["{% block head_extra %}", "<head>", "إضافة CSS أو meta إضافية", "حسب الحاجة"],
    ["{% block body_class %}", "<body>", "كلاسات CSS للـ body", "كل الصفحات"],
    ["{% block header %}", "قبل <main>", "النافبار / الهيدر", "index, courses, course-details (افتراضي) | cart (مُعدّل)"],
    ["{% block main_attrs %}", "<main>", "خصائص إضافية لتاغ main", "حسب الحاجة"],
    ["{% block content %}", "<main>", "المحتوى الرئيسي للصفحة (مطلوب)", "كل الصفحات"],
    ["{% block footer %}", "بعد <main>", "الفوتر", "cart يُعدّله لفوتر مبسط"],
    ["{% block floating_buttons %}", "آخر الـ body", "أزرار عائمة (CTA + scroll top)", "index فقط"],
    ["{% block scripts %}", "قبل </body>", "سكربتات JS خاصة بالصفحة", "كل الصفحات"],
    ["{% block scripts_extra %}", "آخر الـ body", "سكربتات إضافية", "حسب الحاجة"],
]
for i, row in enumerate(blocks, 4):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
style_data(ws, 4, 4 + len(blocks) - 1, 4)
auto_width(ws, 4)

output_path = os.path.join(os.path.dirname(__file__), "02-twig-blocks.xlsx")
wb2.save(output_path)
print(f"Created: {output_path}")


# ============================================================
# FILE 3: Translation Keys
# ============================================================
wb3 = openpyxl.Workbook()
ws = wb3.active
ws.title = "Translation Keys"
ws.sheet_properties.tabColor = "1e3a5f"

add_title(ws, "Translation Keys (مفاتيح الترجمة data-i18n)", 1, 4)

headers = ["Key (المفتاح)", "English (الإنجليزي)", "Arabic (العربي)", "Used In (مُستخدم في)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 4)

translations = [
    # Section header
    ["── Navigation ──", "", "", ""],
    ["nav.brand", "Farabi Academy", "أكاديمية الفارابي", "header, footer"],
    ["nav.aboutUs", "About Us", "من نحن", "header, footer"],
    ["nav.ourCourses", "Our Courses", "دوراتنا", "header, footer"],
    ["nav.successStories", "Success Stories", "قصص النجاح", "header, footer"],
    ["nav.contactUs", "Contact Us", "اتصل بنا", "header, footer"],
    ["nav.registerNow", "Register Now", "سجل الآن", "header"],
    ["nav.login", "Login", "تسجيل الدخول", "header"],
    
    ["── Hero ──", "", "", ""],
    ["hero.title", "Transform Your Future with Excellence", "حوّل مستقبلك مع التميز", "index"],
    ["hero.subtitle", "Join thousands of students who have achieved their dreams...", "انضم إلى آلاف الطلاب الذين حققوا أحلامهم...", "index"],
    ["hero.ctaButton", "Start Now", "ابدأ الآن", "index"],
    
    ["── Courses ──", "", "", ""],
    ["courses.title", "What We Offer", "ما نقدمه", "index"],
    ["courses.subtitle", "Explore our comprehensive range of online courses...", "استكشف مجموعتنا الشاملة من الدورات...", "index"],
    ["courses.viewDetails", "View Details", "عرض التفاصيل", "index"],
    ["courses.course1.title", "Web Development Mastery", "احتراف تطوير الويب", "index"],
    ["courses.course1.description", "Learn HTML, CSS, JavaScript, and React...", "تعلم HTML و CSS و JavaScript و React...", "index"],
    ["courses.course2.title", "Digital Marketing Pro", "التسويق الرقمي المحترف", "index"],
    ["courses.course2.description", "Master SEO, social media marketing...", "أتقن تحسين محركات البحث والتسويق...", "index"],
    ["courses.course3.title", "Graphic Design Fundamentals", "أساسيات التصميم الجرافيكي", "index"],
    ["courses.course3.description", "Create stunning visuals with Adobe...", "أنشئ تصاميم مذهلة باستخدام Adobe...", "index"],
    ["courses.course4.title", "Business Management", "إدارة الأعمال", "index"],
    ["courses.course4.description", "Learn essential business skills...", "تعلم مهارات الأعمال الأساسية...", "index"],
    ["courses.course5.title", "Photography Masterclass", "دورة التصوير الاحترافي", "index"],
    ["courses.course5.description", "From camera basics to professional...", "من أساسيات الكاميرا إلى تقنيات...", "index"],
    ["courses.course6.title", "Data Science & Analytics", "علم البيانات والتحليلات", "index"],
    ["courses.course6.description", "Master Python, SQL, and data visualization...", "أتقن Python و SQL وتصور البيانات...", "index"],
    
    ["── Stats ──", "", "", ""],
    ["stats.students", "Students", "طالب", "index"],
    ["stats.views", "Views", "مشاهدة", "index"],
    ["stats.courses", "Courses", "دورة", "index"],
    ["stats.instructors", "Instructors", "مدرب", "index"],
    ["stats.satisfaction", "Satisfaction", "رضا", "index"],
    
    ["── Reviews ──", "", "", ""],
    ["reviews.title", "What Our Students Say", "ماذا يقول طلابنا", "index"],
    ["reviews.subtitle", "Real stories from real students...", "قصص حقيقية من طلاب حقيقيين...", "index"],
    
    ["── Why Choose Us ──", "", "", ""],
    ["whyChoose.title", "Why Choose Us", "لماذا تختارنا", "index"],
    ["whyChoose.subtitle", "Discover what makes our academy...", "اكتشف ما يجعل أكاديميتنا...", "index"],
    ["whyChoose.feature1Title", "Expert Instructors", "مدربون خبراء", "index"],
    ["whyChoose.feature1Desc", "Learn from industry professionals...", "تعلم من محترفي الصناعة...", "index"],
    ["whyChoose.feature2Title", "Flexible Learning", "تعلم مرن", "index"],
    ["whyChoose.feature2Desc", "Study at your own pace...", "ادرس بالسرعة التي تناسبك...", "index"],
    ["whyChoose.feature3Title", "Lifetime Access", "وصول مدى الحياة", "index"],
    ["whyChoose.feature3Desc", "Get unlimited access...", "احصل على وصول غير محدود...", "index"],
    ["whyChoose.feature4Title", "Certification", "شهادات معترف بها", "index"],
    ["whyChoose.feature4Desc", "Earn recognized certificates...", "احصل على شهادات معترف بها...", "index"],
    ["whyChoose.feature5Title", "Community Support", "دعم المجتمع", "index"],
    ["whyChoose.feature5Desc", "Join a vibrant community...", "انضم إلى مجتمع نابض بالحياة...", "index"],
    ["whyChoose.feature6Title", "Affordable Pricing", "أسعار معقولة", "index"],
    ["whyChoose.feature6Desc", "Quality education shouldn't break...", "التعليم الجيد لا يجب أن يكلف...", "index"],
    
    ["── Contact ──", "", "", ""],
    ["contact.title", "Get In Touch", "تواصل معنا", "index"],
    ["contact.subtitle", "Have questions? We'd love to hear from you...", "لديك أسئلة؟ نحب أن نسمع منك...", "index"],
    ["contact.namePlaceholder", "Your Name", "اسمك", "index"],
    ["contact.emailPlaceholder", "Your Email", "بريدك الإلكتروني", "index"],
    ["contact.messagePlaceholder", "Your Message", "رسالتك", "index"],
    ["contact.sendButton", "Send Message", "إرسال رسالة", "index"],
    ["contact.successMessage", "Thank you! We'll get back to you soon.", "شكراً لك! سنعاود الاتصال بك قريباً.", "index (toast)"],
    
    ["── FAQ ──", "", "", ""],
    ["faq.title", "Frequently Asked Questions", "الأسئلة الشائعة", "index"],
    ["faq.subtitle", "Find answers to common questions...", "اعثر على إجابات للأسئلة الشائعة...", "index"],
    ["faq.q1", "How do I enroll in a course?", "كيف أسجل في دورة؟", "index"],
    ["faq.a1", "Simply click on the 'Register Now' button...", "ببساطة انقر على زر 'سجل الآن'...", "index"],
    ["faq.q2", "Can I access courses on mobile devices?", "هل يمكنني الوصول إلى الدورات على المحمول؟", "index"],
    ["faq.a2", "Yes! Our platform is fully responsive...", "نعم! منصتنا متجاوبة تماماً...", "index"],
    ["faq.q3", "Do you offer refunds?", "هل تقدمون استرداد الأموال؟", "index"],
    ["faq.a3", "We offer a 30-day money-back guarantee...", "نحن نقدم ضمان استرداد الأموال لمدة 30 يوماً...", "index"],
    ["faq.q4", "Are the certificates recognized?", "هل الشهادات معترف بها؟", "index"],
    ["faq.a4", "Yes, our certificates are recognized...", "نعم، شهاداتنا معترف بها...", "index"],
    ["faq.q5", "How long do I have access to a course?", "كم من الوقت لدي للوصول إلى الدورة؟", "index"],
    ["faq.a5", "Once you enroll, you have lifetime access...", "بمجرد تسجيلك، وصول مدى الحياة...", "index"],
    ["faq.q6", "Can I interact with instructors?", "هل يمكنني التفاعل مع المدربين؟", "index"],
    ["faq.a6", "Absolutely! You can ask questions...", "بالتأكيد! يمكنك طرح الأسئلة...", "index"],
    
    ["── Footer ──", "", "", ""],
    ["footer.description", "Empowering learners worldwide...", "تمكين المتعلمين في جميع أنحاء العالم...", "footer"],
    ["footer.quickLinks", "Quick Links", "روابط سريعة", "footer"],
    ["footer.followUs", "Follow Us", "تابعنا", "footer"],
    ["footer.copyright", "© 2026 Farabi Academy. All rights reserved.", "© 2026 أكاديمية الفارابي. جميع الحقوق محفوظة.", "footer"],
    ["floatingCta", "Enroll Now", "سجل الآن", "floating-buttons"],
]

# Apply section header styling
section_fill = PatternFill(start_color='93c572', end_color='93c572', fill_type='solid')
section_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)

row_num = 4
for row_data in translations:
    for j, val in enumerate(row_data, 1):
        ws.cell(row=row_num, column=j, value=val)
    
    # Style section headers differently
    if row_data[0].startswith("──"):
        for j in range(1, 5):
            cell = ws.cell(row=row_num, column=j)
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        # Merge section header
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    row_num += 1

# Style non-section rows
data_font = Font(name='Arial', size=10)
data_align = Alignment(vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
light_fill = PatternFill(start_color='f0f7ec', end_color='f0f7ec', fill_type='solid')

alt = 0
for r in range(4, row_num):
    cell_val = ws.cell(row=r, column=1).value
    if cell_val and cell_val.startswith("──"):
        alt = 0
        continue
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        if alt % 2 == 1:
            cell.fill = light_fill
    alt += 1

# RTL for Arabic column
for r in range(4, row_num):
    cell = ws.cell(row=r, column=3)
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

auto_width(ws, 4, min_width=15, max_width=55)
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 45

output_path = os.path.join(os.path.dirname(__file__), "03-translation-keys.xlsx")
wb3.save(output_path)
print(f"Created: {output_path}")


# ============================================================
# FILE 4: Macros & Components
# ============================================================
wb4 = openpyxl.Workbook()
ws = wb4.active
ws.title = "Macros & Components"
ws.sheet_properties.tabColor = "93c572"

add_title(ws, "Twig Macros & Reusable Components", 1, 4)

headers = ["Macro Name", "File (الملف)", "Parameters (البارامترات)", "Description (الوصف)"]
for i, h in enumerate(headers, 1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, 4)

macros = [
    ["course_card", "pages/index.html.twig", "id, image, price, title_key, desc_key, title_default, desc_default", "كارت كورس في السلايدر بالصفحة الرئيسية"],
    ["review_card", "pages/index.html.twig", "text, image, name, role", "كارت تقييم طالب في سلايدر المراجعات"],
]
for i, row in enumerate(macros, 4):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
style_data(ws, 4, 4 + len(macros) - 1, 4)

# Add file structure
ws.cell(row=8, column=1).value = "Theme File Structure"
ws.cell(row=8, column=1).font = Font(name='Arial', bold=True, size=12, color='1e3a5f')

files = [
    ["File Path", "Description (الوصف)"],
    ["config.json", "إعدادات الثيم (ألوان، معلومات الموقع، روابط السوشيال)"],
    ["lang/en.json", "ملف الترجمة الإنجليزية (60+ مفتاح)"],
    ["lang/ar.json", "ملف الترجمة العربية (60+ مفتاح)"],
    ["layouts/base.html.twig", "القالب الرئيسي (head, meta, scripts, blocks)"],
    ["partials/header.html.twig", "النافبار (desktop + mobile + language toggle)"],
    ["partials/footer.html.twig", "الفوتر (روابط، سوشيال، معلومات تواصل)"],
    ["partials/floating-buttons.html.twig", "الأزرار العائمة (CTA + scroll to top)"],
    ["pages/index.html.twig", "الصفحة الرئيسية (hero, courses, stats, reviews, FAQ)"],
    ["pages/courses.html.twig", "صفحة كل الكورسات (فلاتر، بحث، ترتيب)"],
    ["pages/course-details.html.twig", "صفحة تفاصيل الكورس (فيديو، محتوى، مدرب، مراجعات)"],
    ["pages/cart.html.twig", "صفحة سلة التسوق (عناصر، كوبون، دفع)"],
    ["assets/css/fonts.css", "Google Fonts (Quicksand + Nunito)"],
    ["assets/css/theme.css", "متغيرات CSS، ألوان، أنماط أساسية"],
    ["assets/css/main.css", "أنيميشن، سلايدر، أكورديون، RTL"],
    ["assets/js/utils.js", "دوال مساعدة (scroll, toast, debounce)"],
    ["assets/js/app.js", "التطبيق الرئيسي (i18n, language toggle, nav)"],
    ["assets/js/slider.js", "مكون السلايدر (Vanilla JS)"],
    ["assets/js/counter.js", "أنيميشن العدادات في الإحصائيات"],
    ["assets/js/accordion.js", "مكون الأكورديون"],
    ["assets/js/form.js", "معالج فورم التواصل"],
]

for i, h in enumerate(files[0], 1):
    ws.cell(row=10, column=i, value=h)
style_header(ws, 10, 2)

for i, row in enumerate(files[1:], 11):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)
style_data(ws, 11, 11 + len(files) - 2, 2)

auto_width(ws, 4, min_width=15, max_width=60)

output_path = os.path.join(os.path.dirname(__file__), "04-macros-structure.xlsx")
wb4.save(output_path)
print(f"Created: {output_path}")

print("\n✅ All Excel files generated successfully!")
